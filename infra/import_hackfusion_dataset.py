from __future__ import annotations

import json
import re
from datetime import datetime
from decimal import Decimal
from pathlib import Path

import boto3
from botocore.exceptions import ClientError
from openpyxl import load_workbook

REGION = "us-east-1"
BASE_DIR = Path(r"c:\HackFusion3_Project")
PRODUCTS_XLSX = BASE_DIR / "products-export.xlsx"
ORDERS_XLSX = BASE_DIR / "Consumer Order History 1.xlsx"
PRODUCTS_JSON = BASE_DIR / "products-export.json"
ORDERS_JSON = BASE_DIR / "Consumer Order History 1.json"

MEDICINES_TABLE = "Medicines"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def to_decimal(value: object, default: str = "0") -> Decimal:
    if value is None or value == "":
        return Decimal(default)
    return Decimal(str(value))


def parse_json_object_stream(path: Path):
    if not path.exists():
        return []
    raw = path.read_text(encoding="utf-8", errors="ignore").strip()
    if not raw:
        return []
    wrapped = f"[{raw.rstrip(',')}]"
    try:
        data = json.loads(wrapped)
        return data if isinstance(data, list) else []
    except Exception:
        # fallback: parse individual {...} chunks
        chunks = re.findall(r"\{.*?\}", raw, flags=re.DOTALL)
        out = []
        for c in chunks:
            try:
                out.append(json.loads(c))
            except Exception:
                continue
        return out


def load_products_xlsx(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb["Products"] if "Products" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [normalize_text(c) for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(header, row))
        rows.append(
            {
                "product_id": normalize_text(rec.get("product id")),
                "medicine_name": normalize_text(rec.get("product name")),
                "pzn": normalize_text(rec.get("pzn")),
                "price": to_decimal(rec.get("price rec"), "0"),
                "package_size": normalize_text(rec.get("package size")),
                "description": normalize_text(rec.get("descriptions")),
            }
        )
    return rows


def load_products_json(path: Path) -> list[dict]:
    rows = []
    for rec in parse_json_object_stream(path):
        if not isinstance(rec, dict):
            continue
        name = normalize_text(rec.get("product name"))
        if not name:
            continue
        rows.append(
            {
                "product_id": normalize_text(rec.get("product id")),
                "medicine_name": name,
                "pzn": normalize_text(rec.get("pzn")),
                "price": to_decimal(rec.get("price rec"), "0"),
                "package_size": normalize_text(rec.get("package size")),
                "description": normalize_text(rec.get("descriptions")),
            }
        )
    return rows


def find_order_header_row(ws) -> int:
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
        first = normalize_text(row[0]) if row else ""
        if first.lower() == "patient id":
            return i
    raise RuntimeError("Could not find order header row (Patient ID)")


def load_orders_xlsx(path: Path) -> list[dict]:
    if not path.exists():
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb[wb.sheetnames[0]]
    header_row = find_order_header_row(ws)
    header = [normalize_text(c) for c in next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))]

    rows = []
    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if not any(row):
            continue
        rec = dict(zip(header, row))
        patient_id = normalize_text(rec.get("Patient ID"))
        if not patient_id:
            continue

        purchase_date = rec.get("Purchase Date")
        if isinstance(purchase_date, datetime):
            purchase_date = purchase_date.date().isoformat()
        else:
            purchase_date = normalize_text(purchase_date)

        rows.append(
            {
                "patient_id": patient_id,
                "patient_age": int(rec.get("Patient Age") or 0),
                "patient_gender": normalize_text(rec.get("Patient Gender")),
                "purchase_date": purchase_date,
                "medicine_name": normalize_text(rec.get("Product Name")),
                "quantity": int(rec.get("Quantity") or 1),
                "total_price": to_decimal(rec.get("Total Price (EUR)"), "0"),
                "dosage_frequency": normalize_text(rec.get("Dosage Frequency")),
                "prescription_required": normalize_text(rec.get("Prescription Required")).lower() in {"yes", "y", "true", "1"},
            }
        )
    return rows


def load_orders_json(path: Path) -> list[dict]:
    rows = []
    for rec in parse_json_object_stream(path):
        if not isinstance(rec, dict):
            continue
        patient_id = normalize_text(rec.get("Consumer Order History - Pharmaceutical Products"))
        if not patient_id or not re.fullmatch(r"PAT\d+", patient_id.upper()):
            continue

        rows.append(
            {
                "patient_id": patient_id,
                "patient_age": int(rec.get("Column2") or 0),
                "patient_gender": normalize_text(rec.get("Column3")),
                "purchase_date": normalize_text(rec.get("Column4")),
                "medicine_name": normalize_text(rec.get("Column5")),
                "quantity": int(rec.get("Column6") or 1),
                "total_price": to_decimal(rec.get("Column7"), "0"),
                "dosage_frequency": normalize_text(rec.get("Column8")),
                "prescription_required": normalize_text(rec.get("Column9")).lower() in {"yes", "y", "true", "1"},
            }
        )
    return rows


def dedupe_products(products: list[dict]) -> list[dict]:
    merged = {}
    for p in products:
        key = p.get("medicine_name")
        if not key:
            continue
        if key not in merged:
            merged[key] = p
        else:
            for k, v in p.items():
                if (not normalize_text(merged[key].get(k))) and normalize_text(v):
                    merged[key][k] = v
    return list(merged.values())


def dedupe_orders(orders: list[dict]) -> list[dict]:
    seen = set()
    out = []
    for o in orders:
        key = (
            o.get("patient_id"),
            o.get("purchase_date"),
            o.get("medicine_name"),
            int(o.get("quantity", 0)),
            str(o.get("total_price")),
        )
        if key in seen:
            continue
        seen.add(key)
        out.append(o)
    return out


def ensure_product_flags(products: list[dict], orders: list[dict]) -> list[dict]:
    presc_map = {}
    sold_qty = {}
    for row in orders:
        name = row["medicine_name"]
        if not name:
            continue
        presc_map[name] = presc_map.get(name, False) or bool(row["prescription_required"])
        sold_qty[name] = sold_qty.get(name, 0) + int(row["quantity"])

    for p in products:
        sold = sold_qty.get(p["medicine_name"], 0)
        p["stock"] = max(50, sold * 3)
        p["requires_prescription"] = presc_map.get(p["medicine_name"], False)
    return products


def upsert_medicines(table, products: list[dict]) -> int:
    count = 0
    for p in products:
        table.put_item(
            Item={
                "medicine_name": p["medicine_name"],
                "product_id": p.get("product_id", ""),
                "pzn": p.get("pzn", ""),
                "price": p.get("price", Decimal("0")),
                "package_size": p.get("package_size", ""),
                "description": p.get("description", ""),
                "stock": int(p.get("stock", 0)),
                "requires_prescription": bool(p.get("requires_prescription", False)),
                "source": "hackfusion_dataset_all",
            }
        )
        count += 1
    return count


def main():
    products = dedupe_products(load_products_xlsx(PRODUCTS_XLSX) + load_products_json(PRODUCTS_JSON))
    orders = dedupe_orders(load_orders_xlsx(ORDERS_XLSX) + load_orders_json(ORDERS_JSON))

    if not products:
        raise RuntimeError("No product data found in provided XLSX/JSON sources")
    if not orders:
        raise RuntimeError("No order data found in provided XLSX/JSON sources")

    products = ensure_product_flags(products, orders)

    dynamodb = boto3.resource("dynamodb", region_name=REGION)
    medicines = dynamodb.Table(MEDICINES_TABLE)
    try:
        medicines.meta.client.describe_table(TableName=MEDICINES_TABLE)
    except ClientError as e:
        raise RuntimeError("Medicines table must exist. Run infra/setup_dynamodb.py first.") from e

    m_count = upsert_medicines(medicines, products)

    print(f"Imported medicines (all sources): {m_count}")
    print("Skipped Orders import by design (dataset orders are not written to DynamoDB)")
    print("Dataset import complete (Medicines only)")


if __name__ == "__main__":
    main()
